"""
Chạy local:  uvicorn main:app --host 0.0.0.0 --port 8000
"""

from fastapi import FastAPI, File, UploadFile, Form, Query
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from collections import deque, Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
import asyncio
import cv2
import numpy as np
import base64
import os
import glob
import time
import io
import calendar
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from recognizer import FaceRecognizer, BASE_DIR, DB_PATH
from plc_light import PLCLight

executor = ThreadPoolExecutor(max_workers=3)

app = FastAPI(title="YOLO26 + ArcFace Face Recognition")

#   YOLO26 (best.pt)      -> phát hiện và khoanh mặt
#   ArcFace              -> embedding 512 chiều cho mỗi mặt
#   Cosine với face_db.pt -> ra tên, hoặc "người lạ" nếu không đủ gần


CONF_DETECT = 0.25   # ngưỡng YOLO bắt mặt (để thấp cho đỡ bỏ sót)
THRESHOLD   = 0.28   # ngưỡng cosine (trung bình top-5): >= coi là người quen.
                     # Đo trên test: đúng 100% ở khoảng 0.25-0.28. Xem eval_recognition.py.

# Số frame gần nhất của cùng một mặt dùng để làm mượt điểm số. Nhỏ thì nhận ra
# nhanh hơn với người đi qua, lớn thì ổn định hơn với người đứng im.
SMOOTH_WINDOW = 5

# Danh sách người đã enroll sẵn (có thể thêm người mới qua /capture, không cần train).
KNOWN_PEOPLE = ["nghia", "quan", "son", "tri", "tui"]

# Nhật ký ra vào, dùng cho sheet Tuần/Tháng khi export Excel.
# Mỗi entry: {dt, type, person, role, duration_min, ks_present, sv_present, total_ks, total_sv}
# Check-in/check-out tách theo camera (xem CAM_ROLE, _log_attendance).
_attendance_log: list[dict] = []
_in_room: dict[str, datetime] = {}   # name -> thời điểm vào phòng

# Nhật ký quét mặt, dùng cho sheet Ngày. Mỗi entry: {dt, person, role, status, motion, cam}
#   status: OK   = nhận ra người quen (mỗi track chỉ ghi 1 lần)
#           FAIL = đang là người quen X bỗng thành người lạ (nhận diện lỗi)
#           LẠ   = người lạ hoàn toàn
#   motion: "Đứng im" | "Di chuyển"
_scan_log: list[dict] = []

# 2 camera chạy 2 luồng độc lập, mỗi camera có state tracking riêng.
CAMERAS = {"cam1": "UGreen 2K", "cam2": "Rapoo C260"}   # id -> tên hiển thị

# Siết nhận diện riêng cho từng camera. Rapoo C260 (cam2) cho ảnh mềm hơn UGreen 2K
# nên dễ gán nhầm tên. Với cam2:
#   thr_bump : cộng thêm vào ngưỡng cosine người quen.
#   margin   : top-1 phải hơn top-2 ít nhất chừng này, nếu không thì coi frame là
#              chưa chắc và để bộ làm mượt tự quyết.
# cam1 để 0/0 nên giữ nguyên hành vi cũ. Đây là giá trị khởi đầu, nên soi lại Excel
# người lạ/fail rồi chỉnh nếu còn nhầm hoặc bị miss.
CAM_TUNE = {
    "cam1": {"thr_bump": 0.00, "margin": 0.00},
    "cam2": {"thr_bump": 0.07, "margin": 0.05},
}

# UGreen (cam1) đặt ở lối vào, Rapoo (cam2) đặt ở lối ra. Check-in/out gán thẳng
# theo camera nhận ra người.
CAM_ROLE = {"cam1": "vào", "cam2": "ra"}

# Lọc mặt mờ (motion blur) trước khi chạy ArcFace. Ảnh mờ vẫn ra được embedding
# nhưng kém tin cậy nên dễ nhận nhầm người quen này thành người quen khác. Khung mờ
# thì bỏ qua, không tính phiếu, track vẫn bám vị trí và chờ khung nét hơn. Đo bằng
# phương sai Laplacian trên vùng mặt resize 100x100. Rapoo (cam2) vốn mềm hơn nên
# ngưỡng thấp hơn cam1.
BLUR_THR = {"cam1": 60.0, "cam2": 35.0}
SKIP_VOTE = "__SKIP__"   # best_name khi khung bị lọc mờ: không tính là phiếu

# Bù sáng và độ nét cho camera thiếu AF/HDR. Rapoo không có HDR (dễ cháy/tối khi
# sáng-tối lệch) và fixed-focus (ảnh mềm). Xử lý trên ảnh gốc trước khi đưa vào YOLO
# và ArcFace:
#   CLAHE (kênh L của LAB) -> cân sáng cục bộ, bù thiếu HDR.
#   Unsharp mask nhẹ       -> bù mềm do fixed-focus.
# cam1 đã có AF + HDR phần cứng nên để False.
CAM_ENHANCE = {"cam1": False, "cam2": True}
_clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))

def _enhance(img, cam):
    if not CAM_ENHANCE.get(cam, False):
        return img
    l, a, b = cv2.split(cv2.cvtColor(img, cv2.COLOR_BGR2LAB))
    l = _clahe.apply(l)
    img = cv2.cvtColor(cv2.merge((l, a, b)), cv2.COLOR_LAB2BGR)
    blur = cv2.GaussianBlur(img, (0, 0), 3)
    return cv2.addWeighted(img, 1.35, blur, -0.35, 0)   # unsharp mask nhẹ

class CamState:
    def __init__(self):
        self.tracks = []              # track của riêng camera này
        self.stable_start = {}        # key -> lúc bắt đầu đứng im
        self.stable_seen  = {}        # key -> lần thấy đứng im gần nhất
        self.stable_logged = set()    # key đã ghi đứng im (1 lần/người/camera)

_cam_states: dict = defaultdict(CamState)

# Đèn tháp PLC Mitsubishi Q/iQ-R qua MC Protocol. Chỉnh IP/port/địa chỉ Y cho khớp
# tủ điện thật trước khi chạy thật.
PLC_IP      = "127.0.0.1"          # tạm trỏ vào plc_simulator.py để test/demo
PLC_PORT    = 5007                 # khớp --port của plc_simulator.py
PLC_TYPE    = "Q"                  # "Q" hoặc "iQ-R"
PLC_COILS   = {"green": "Y0", "yellow": "Y1", "red": "Y2"}   # TODO: đổi cho đúng địa chỉ Y thật
PLC_PULSE_S = 2.5                  # thời gian đèn sáng mỗi lần báo (giây)

plc = PLCLight(PLC_IP, PLC_PORT, PLC_COILS, plctype=PLC_TYPE, pulse_sec=PLC_PULSE_S)


def _log_scan(person: str, status: str, cam: str, motion: str = "Đứng im"):
    _scan_log.append({"dt": datetime.now(), "person": person,
                      "role": rec.db_roles.get(person, "kỹ sư"),
                      "status": status, "motion": motion,
                      "cam": CAMERAS.get(cam, cam)})
    if status == "OK":
        _log_attendance(person, cam)   # OK -> check-in/out theo CAM_ROLE
    # Báo đèn theo trạng thái nhận diện: OK -> xanh, FAIL -> vàng, LẠ -> đỏ.
    plc.signal({"OK": "green", "FAIL": "yellow", "LẠ": "red"}.get(status, "red"))

print("[main] Khởi tạo YOLO26 + ArcFace...")
rec = FaceRecognizer(det_conf=CONF_DETECT)   # YOLO detect mặt, ArcFace nhận diện tên

# Chế độ nhẹ khi máy không có GPU: hạ imgsz YOLO 512 -> 416 cho nhanh. YOLO chỉ tìm
# vị trí mặt (ArcFace tự cắt và căn lại từ ảnh gốc) nên không giảm độ chính xác.
LIGHT_MODE = (rec.device == "cpu")
if LIGHT_MODE:
    rec.det_imgsz = 416
TRACK_IMGSZ = 256 if LIGHT_MODE else 320   # imgsz YOLO cho /track (chỉ bám khung)

print(f"[main] Device YOLO : {rec.device}   {'(chế độ nhẹ - CPU)' if LIGHT_MODE else ''}")
print(f"[main] DB          : {DB_PATH}  ({len(rec.db_names)} người: {rec.db_names})")
print(f"[main] Ngưỡng      : detect={CONF_DETECT}  cosine={THRESHOLD}")

# Warm-up để lần nhận diện đầu không bị chậm
rec.recognize(np.zeros((320, 320, 3), dtype=np.uint8), threshold=THRESHOLD)
print("[main] Warm-up done")


# Bộ theo dõi và làm mượt điểm số, chống nhấp nháy "người lạ". Mỗi mặt ghép với một
# track theo vị trí; track giữ lịch sử (tên, điểm) vài frame và quyết định theo điểm
# trung bình của danh tính nổi trội trong cửa sổ. Track state nằm trong CamState,
# mỗi camera một bộ riêng (xem _cam_states).

HYSTERESIS  = 0.04    # đã nhận tên thì hạ ngưỡng giữ tên 0.04 (khó rớt sang "lạ")
MIN_PRESENCE = 0.5    # danh tính phải chiếm >= 50% cửa sổ mới được nhận
MATCH_FRAC  = 0.18     # ngưỡng ghép track = 18% đường chéo khung
MOVE_SPEED_FRAC = 0.6  # tốc độ tâm mặt > 60% bề rộng mặt mỗi giây thì coi là di chuyển
STABLE_SECONDS = 2.0   # đứng im đủ 2 giây thì ghi log 1 lần
STABLE_GAP  = 4.0      # gián đoạn thấy > 4s thì coi là lượt đứng mới (phải lớn hơn nhịp
                       # nhận diện trên máy chậm, không thì đồng hồ đứng im bị reset hoài)
STRANGER_MIN = 4       # phải "lạ" liên tục >= 4 frame mới ghi người lạ
TRACK_TTL   = 1.5      # track không thấy > 1.5 giây thì xoá
REFRESH_SEC = 3.0      # đã nhận ra người quen thì dùng lại tên, chỉ chạy ArcFace lại sau 3s
REACQUIRE_GAP = 0.6    # ghép lại một track sau khi mất dấu > 0.6s (nhưng < TRACK_TTL nên
                       # chưa bị xoá): coi như có thể là người khác đứng đúng chỗ đó, xoá
                       # danh tính cũ và nhận diện lại từ đầu. Người đi liên tục ghép mỗi
                       # ~100ms qua /track nên gap luôn < 0.6s, không bị reset.

def _reacquire_if_stale(t, now):
    """Gọi trước khi cập nhật seen cho một track vừa được ghép lại (xem REACQUIRE_GAP)."""
    if (now - t["seen"]).total_seconds() > REACQUIRE_GAP:
        t["hist"].clear(); t["label"] = "LẠ"; t["logged_label"] = ""
        t["ever_known"] = False; t["arc_time"] = None; t["arc_score"] = 0.5

def _is_moving(pos, wsize):
    """pos: deque (cx, cy, dt). Đo tốc độ theo giây (độc lập FPS) bằng cách so vị trí
    trung bình nửa đầu và nửa sau cửa sổ, để một điểm nhiễu không làm đứng im thành
    di chuyển."""
    if len(pos) < 3 or wsize <= 0:
        return False
    h = len(pos) // 2
    first, last = list(pos)[:h], list(pos)[-h:]
    fx = sum(p[0] for p in first) / h; fy = sum(p[1] for p in first) / h
    lx = sum(p[0] for p in last) / h;  ly = sum(p[1] for p in last) / h
    dt = (last[-1][2] - first[0][2]).total_seconds()
    if dt < 0.15:
        return False
    disp = ((lx - fx) ** 2 + (ly - fy) ** 2) ** 0.5
    return (disp / dt) > MOVE_SPEED_FRAC * wsize

def _match_thr():
    # ngưỡng ghép trong hệ toạ độ chuẩn hoá [0,1] (đường chéo = căn 2)
    return MATCH_FRAC * (2 ** 0.5)

def _new_track(cx, cy, now):
    return {"cx": cx, "cy": cy, "hist": deque(maxlen=SMOOTH_WINDOW),
            "seen": now, "label": "LẠ", "logged_label": "",
            "pos": deque(maxlen=SMOOTH_WINDOW), "moving": False, "ever_known": False,
            "arc_time": None, "arc_score": 0.5}   # cache: lần chạy ArcFace gần nhất và điểm

def _nearest_track(tracks, cxn, cyn, thr):
    """Track gần nhất (chỉ đọc, không chiếm), để quyết định có cần chạy ArcFace không."""
    best, bd = None, thr
    for t in tracks:
        d = ((t["cx"] - cxn) ** 2 + (t["cy"] - cyn) ** 2) ** 0.5
        if d < bd:
            bd, best = d, t
    return best

def _match_track_idx(tracks, cx, cy, thr, used):
    """Ghép (cx,cy) với track gần nhất chưa dùng trong `tracks`. Trả index hoặc None."""
    bt, bd = None, thr
    for ti, t in enumerate(tracks):
        if ti in used:
            continue
        d = ((t["cx"] - cx) ** 2 + (t["cy"] - cy) ** 2) ** 0.5
        if d < bd:
            bd, bt = d, ti
    return bt

def _prune_tracks(st, now):
    """Dọn track cũ của một camera theo thời gian (độc lập FPS)."""
    st.tracks = [t for t in st.tracks if (now - t["seen"]).total_seconds() <= TRACK_TTL]

def smooth_labels(faces, W, H, threshold, st, cam):
    """faces -> list (label, disp_score) đã làm mượt, dùng track state `st` của camera
    `cam`. Ghi nhật ký kèm tên camera."""
    thr = _match_thr()
    now = datetime.now()
    _prune_tracks(st, now)   # dọn track cũ trước khi ghép
    tracks = st.tracks
    used, out = set(), []
    for f in faces:
        x1, y1, x2, y2 = f["box"]
        # toạ độ chuẩn hoá [0,1], chung hệ với /track
        cxn, cyn = (x1 + x2) / 2 / W, (y1 + y2) / 2 / H
        wsn = (x2 - x1) / W
        bn, bs = f.get("best_name"), float(f.get("score", 0.0))

        bt = _match_track_idx(tracks, cxn, cyn, thr, used)
        if bt is None:
            t = _new_track(cxn, cyn, now)
            tracks.append(t); used.add(len(tracks) - 1)
        else:
            t = tracks[bt]
            _reacquire_if_stale(t, now)   # gián đoạn dài: có thể người khác, xoá danh tính cũ
            t["cx"], t["cy"], t["seen"] = cxn, cyn, now
            used.add(bt)
        if bn != SKIP_VOTE:   # khung mờ thì không tính phiếu, giữ nguyên hist (xem BLUR_THR)
            t["hist"].append((bn, bs))

        # đứng im hay di chuyển
        t["pos"].append((cxn, cyn, now))
        t["moving"] = _is_moving(t["pos"], wsn)

        # gộp điểm theo từng tên trong cửa sổ rồi chọn danh tính nổi trội
        by = defaultdict(list)
        for n, s in t["hist"]:
            if n is not None:
                by[n].append(s)
        if by:
            dom = max(by, key=lambda n: (len(by[n]), sum(by[n]) / len(by[n])))
            mean_s   = sum(by[dom]) / len(by[dom])
            presence = len(by[dom]) / len(t["hist"])
            # track đang mang chính tên này thì ngưỡng giữ thấp hơn
            keep_thr = threshold - HYSTERESIS if t["label"] == dom else threshold
            if presence >= MIN_PRESENCE and mean_s >= keep_thr:
                t["label"] = dom
            else:
                t["label"] = "LẠ"
            disp = mean_s
        else:
            t["label"] = "LẠ"
            disp = bs
        out.append((t["label"], round(disp, 3)))

        # Ghi nhật ký quét:
        #   đứng im  : một danh tính đứng đủ STABLE_SECONDS thì ghi 1 lần, lặp lại thì bỏ qua.
        #   di chuyển: ghi OK/FAIL mỗi lần đổi trạng thái.
        old = t.get("logged_label", "")
        lbl = t["label"]
        if lbl != "LẠ":
            t["ever_known"] = True               # track này đã từng là người quen
        key = lbl if lbl != "LẠ" else "Người lạ"

        if t["moving"]:
            if lbl != "LẠ" and len(t["hist"]) >= 1:
                _log_scan(lbl, "OK", cam, "Di chuyển"); t["logged_label"] = lbl
            elif lbl == "LẠ" and old not in ("", "LẠ"):
                _log_scan(old, "FAIL", cam, "Di chuyển"); t["logged_label"] = lbl
            elif lbl == "LẠ" and old != "LẠ" and not t["ever_known"] and len(t["hist"]) >= STRANGER_MIN:
                _log_scan("Người lạ", "LẠ", cam, "Di chuyển"); t["logged_label"] = lbl
            # đang di chuyển thì reset đồng hồ đứng im (giữ cờ đã-ghi)
            st.stable_start.pop(key, None); st.stable_seen.pop(key, None)
        else:
            # đứng im: mỗi người chỉ ghi 1 lần trên camera này
            ok_stable = (lbl != "LẠ" and len(t["hist"]) >= 1)
            la_stable = (lbl == "LẠ" and not t["ever_known"] and len(t["hist"]) >= STRANGER_MIN)
            if (ok_stable or la_stable) and key not in st.stable_logged:
                prev = st.stable_seen.get(key)
                if prev is None or (now - prev).total_seconds() > STABLE_GAP:
                    st.stable_start[key] = now          # bắt đầu tính giờ đứng
                st.stable_seen[key] = now
                if (now - st.stable_start[key]).total_seconds() >= STABLE_SECONDS:
                    _log_scan(key, "LẠ" if key == "Người lạ" else "OK", cam, "Đứng im")
                    st.stable_logged.add(key)
                    t["logged_label"] = lbl

    return out

@app.get("/", response_class=HTMLResponse)
def read_root():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()


@app.get("/health")
def health():
    """Kiểm tra server còn sống — dùng cho Docker HEALTHCHECK."""
    return {
        "status": "ok",
        "arch": "YOLO26 detect + ArcFace embedding",
        "device": rec.device,
        "light_mode": LIGHT_MODE,
        "people": rec.db_names,
        "db_size": 0 if rec.db_embs is None else int(rec.db_embs.shape[0]),
        "threshold": THRESHOLD,
        "plc_connected": plc.connected,
    }


@app.get("/plc-status")
def plc_status():
    """Trạng thái đèn tháp hiện tại, đọc thẳng từ PLCLight. Web UI poll endpoint này
    để hiển thị đúng màu đang sáng, kể cả khi PLC thật chưa đấu nối."""
    return {"color": plc.current, "connected": plc.connected}


@app.get("/plc-test")
def plc_test(color: str = Query("green")):
    """Bấm thử đèn tháp từ trình duyệt để kiểm tra đấu dây/địa chỉ Y:
    GET /plc-test?color=green | yellow | red"""
    if color not in PLC_COILS:
        return JSONResponse({"ok": False, "msg": f"Màu '{color}' không hợp lệ (green/yellow/red)"},
                            status_code=400)
    plc.signal(color)
    return {"ok": True, "color": color, "device": PLC_COILS[color]}


@app.post("/process-frame")
async def process_frame(
    file: UploadFile = File(...),
    conf_detect: float = Form(CONF_DETECT),   # ngưỡng YOLO bắt mặt (web gửi lên)
    conf_known: float = Form(THRESHOLD),      # ngưỡng cosine người quen/lạ (web gửi lên)
    cam: str = Form("cam1"),                  # camera nào gửi frame (state tracking riêng)
):
    """
    Nhận frame webcam/video -> YOLO khoanh mặt -> ArcFace embedding -> cosine với DB.
    Trả về toạ độ khung và danh sách nhận diện. Ngưỡng chỉnh trực tiếp từ web.
    """
    contents = await file.read()
    nparr = np.frombuffer(contents, np.uint8)
    img   = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None:
        return JSONResponse({"status": "error", "message": "Invalid image"}, status_code=400)
    img = _enhance(img, cam)   # bù sáng/độ nét cho cam thiếu AF/HDR (xem CAM_ENHANCE)

    # kẹp giá trị hợp lệ phòng web gửi số vô lý
    conf_detect = min(max(conf_detect, 0.05), 0.9)
    threshold   = min(max(conf_known, 0.05), 0.95)
    # siết riêng theo camera (xem CAM_TUNE)
    tune     = CAM_TUNE.get(cam, CAM_TUNE["cam1"])
    threshold = min(threshold + tune["thr_bump"], 0.95)
    margin    = tune["margin"]

    st = _cam_states[cam]                 # state tracking riêng của camera này
    rec.det_conf = conf_detect
    t0 = time.perf_counter()
    loop = asyncio.get_event_loop()
    H, W = img.shape[:2]

    # Nhận diện có cache để tiết kiệm CPU:
    #   1) YOLO tìm mọi mặt.
    #   2) Mặt đã nhận ra chắc và còn hạn thì dùng lại tên, không chạy ArcFace.
    #   3) ArcFace chỉ chạy cho mặt mới / chưa chắc / quá hạn refresh.
    boxes = await loop.run_in_executor(executor, lambda: rec.detect(img))
    now_dt = datetime.now()
    thr_m = _match_thr()
    blur_thr = BLUR_THR.get(cam, BLUR_THR["cam1"])
    faces = [None] * len(boxes)
    trk_of, need = [], []
    for i, b in enumerate(boxes):
        x1, y1, x2, y2 = b[:4]
        cxn, cyn = (x1 + x2) / 2 / W, (y1 + y2) / 2 / H
        tr = _nearest_track(st.tracks, cxn, cyn, thr_m)
        trk_of.append(tr)
        if (tr is not None and tr["label"] != "LẠ" and tr["arc_time"] is not None
                and (now_dt - tr["arc_time"]).total_seconds() < REFRESH_SEC):
            faces[i] = {"box": (x1, y1, x2, y2), "best_name": tr["label"],
                        "score": tr["arc_score"], "runner": None}   # dùng lại tên, bỏ ArcFace
        elif rec.is_blurry(img, (x1, y1, x2, y2), blur_thr):
            # khung mờ: bỏ ArcFace, không tính phiếu (xem BLUR_THR/SKIP_VOTE)
            faces[i] = {"box": (x1, y1, x2, y2), "best_name": SKIP_VOTE, "score": 0.0, "runner": None}
        else:
            need.append(i)
    if need:
        need_boxes = [boxes[i][:4] for i in need]
        emb, keep = await loop.run_in_executor(executor, lambda: rec.embed(img, need_boxes))
        got = {}
        for j, k in enumerate(keep):
            ranked = rec._rank(emb[j])
            if ranked:
                top_name, top_score = ranked[0]
                run_score = ranked[1][1] if len(ranked) > 1 else 0.0
                runner = ({"name": ranked[1][0], "score": round(ranked[1][1], 3)}
                          if len(ranked) > 1 else None)
                # top-1 sát top-2 nghĩa là đang lẫn giữa 2 người quen: bỏ tên (best_name=None)
                # để frame này thành phiếu chưa chắc. Mặt đang di chuyển vốn mờ hơn nên
                # top1/top2 sát nhau hơn, giảm nửa margin khi track đang moving.
                m_tr = trk_of[need[k]]
                m = margin * 0.5 if (m_tr is not None and m_tr.get("moving")) else margin
                name = None if (top_score - run_score) < m else top_name
                got[k] = (name, round(top_score, 3), runner)
        for pos, i in enumerate(need):
            x1, y1, x2, y2 = boxes[i][:4]
            info = got.get(pos)
            if info is None:
                faces[i] = {"box": (x1, y1, x2, y2), "best_name": None, "score": 0.0, "runner": None}
            else:
                bn, bs, runner = info
                faces[i] = {"box": (x1, y1, x2, y2), "best_name": bn, "score": bs, "runner": runner}

                if trk_of[i] is not None:
                    trk_of[i]["arc_time"] = now_dt
                    trk_of[i]["arc_score"] = bs
    faces = [f for f in faces if f is not None]
    infer_ms = (time.perf_counter() - t0) * 1000

    n0 = len(_scan_log)
    voted = smooth_labels(faces, W, H, threshold, st, cam)   # làm mượt và ghi log theo camera
    # sự kiện quét mới sinh ra ở frame này, gửi về client để hiện live
    scan_events = [{"time": e["dt"].strftime("%H:%M:%S"), "person": e["person"],
                    "status": e["status"], "motion": e["motion"], "cam": e["cam"]}
                   for e in _scan_log[n0:]]

    stranger = False
    known = []
    detections = []

    # Chỉ trả toạ độ khung, client tự vẽ overlay lên video (nhẹ băng thông, nhanh hơn).
    for f, (vname, vscore) in zip(faces, voted):
        x1, y1, x2, y2 = f["box"]
        runner = f.get("runner")
        box = [int(x1), int(y1), int(x2), int(y2)]
        if vname == "LẠ":
            stranger = True
            detections.append({"name": "Người lạ", "conf": vscore, "stranger": True,
                               "runner": runner, "box": box})
        else:
            known.append(vname)
            detections.append({"name": vname, "conf": vscore, "stranger": False,
                               "runner": runner, "box": box})

    known_unique = list(dict.fromkeys(known))   # bỏ trùng, giữ thứ tự
    # người lạ lên đầu, rồi theo cosine giảm dần
    detections.sort(key=lambda d: (not d["stranger"], -d["conf"]))
    return {
        "img_w": W, "img_h": H,   # kích thước frame để client scale toạ độ khung
        "stranger": stranger,
        "known": known_unique,
        "detections": detections,
        "count": len(faces),
        "infer_ms": round(infer_ms, 1),
        "scan_events": scan_events,   # sự kiện OK/FAIL/LẠ mới (hiện live trên UI)
    }


@app.post("/track")
async def track_faces(
    file: UploadFile = File(...),
    conf_detect: float = Form(CONF_DETECT),
    cam: str = Form("cam1"),                  # camera nào (state tracking riêng)
):
    """Chỉ chạy YOLO (~50ms) rồi trả toạ độ khung bám sát thời gian thực. Tên/điểm lấy
    từ track gần nhất (do /process-frame cập nhật) nên không phải chờ ArcFace. Client
    gọi liên tục cho khung mượt và biến mất nhanh khi rời khung hình."""
    contents = await file.read()
    img = cv2.imdecode(np.frombuffer(contents, np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        return JSONResponse({"status": "error", "message": "Invalid image"}, status_code=400)

    rec.det_conf = min(max(conf_detect, 0.05), 0.9)
    loop = asyncio.get_event_loop()
    H, W = img.shape[:2]
    # /track chạy ~10fps chỉ để bám khung nên không áp _enhance (đỡ tốn CPU ở vòng lặp
    # tần suất cao). Nhận diện thật nằm ở /process-frame. imgsz nhỏ cho nhanh trên CPU.
    boxes = await loop.run_in_executor(executor, lambda: rec.detect(img, imgsz=TRACK_IMGSZ))

    st = _cam_states[cam]                 # state tracking riêng của camera này
    match_thr = _match_thr()
    now = datetime.now()
    _prune_tracks(st, now)   # dọn track cũ trước khi ghép
    tracks = st.tracks
    used, dets = set(), []
    for (x1, y1, x2, y2, conf) in boxes:
        # toạ độ chuẩn hoá [0,1], chung hệ với /process-frame
        cxn, cyn = (x1 + x2) / 2 / W, (y1 + y2) / 2 / H
        wsn = (x2 - x1) / W
        bi = _match_track_idx(tracks, cxn, cyn, match_thr, used)
        if bi is None:
            # tạo track mới, chờ /process-frame gán tên
            t = _new_track(cxn, cyn, now)
            tracks.append(t); bi = len(tracks) - 1
        else:
            t = tracks[bi]
            _reacquire_if_stale(t, now)   # gián đoạn dài: có thể người khác, xoá danh tính cũ
            t["cx"], t["cy"], t["seen"] = cxn, cyn, now
        used.add(bi)
        # góp dữ liệu chuyển động ở nhịp nhanh của /track để đo đứng/đi chính xác hơn
        t["pos"].append((cxn, cyn, now))
        t["moving"] = _is_moving(t["pos"], wsn)

        lab = t["label"]
        if not t["hist"]:                 
            name, score, stranger, pending = "…", 0.0, False, True
        elif lab == "LẠ":
            name, score, stranger, pending = "Người lạ", t["hist"][-1][1], True, False
        else:
            name, score, stranger, pending = lab, t["hist"][-1][1], False, False
        dets.append({"name": name, "conf": round(float(score), 3), "stranger": stranger,
                     "pending": pending, "moving": t["moving"],
                     "box": [int(x1), int(y1), int(x2), int(y2)]})

    dets.sort(key=lambda d: (not d["stranger"], -d["conf"]))
    return {"img_w": W, "img_h": H, "detections": dets, "count": len(boxes)}


@app.post("/capture")
async def capture(file: UploadFile = File(...), person: str = Form(...), role: str = Form("kỹ sư")):
    """Thêm 1 mẫu ảnh vào DB (không cần train lại). Ghi ra face_db.pt."""
    person = (person or "").strip()
    if not person:
        return JSONResponse({"ok": False, "msg": "Chưa nhập tên người"}, status_code=200)

    contents = await file.read()
    img = cv2.imdecode(np.frombuffer(contents, np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        return JSONResponse({"ok": False, "msg": "Ảnh lỗi"}, status_code=400)

    vec = rec.embed_largest(img)   # YOLO khoanh mặt to nhất rồi ArcFace embedding
    if vec is None:
        return JSONResponse({"ok": False, "msg": "Không thấy/căn được khuôn mặt — đưa mặt gần và thẳng hơn"},
                            status_code=200)

    is_new = person not in rec.db_names
    rec.add_embedding(person, vec, role=role)   # tạo người mới nếu chưa có
    rec.save_db()                    # ghi ra face_db.pt

    count = int((rec.db_owner == rec.db_names.index(person)).sum())
    total = rec.db_embs.shape[0]
    return {"ok": True, "person": person, "count": count, "total": total,
            "is_new": is_new, "people": rec.people_summary()}


@app.get("/people")
def people():
    return {"people": rec.people_summary(),
            "total": 0 if rec.db_embs is None else int(rec.db_embs.shape[0])}


@app.post("/update-role")
async def update_role(person: str = Form(...), role: str = Form(...)):
    person = (person or "").strip()
    if person not in rec.db_names:
        return JSONResponse({"ok": False, "msg": f"Không tìm thấy '{person}'"}, status_code=200)
    rec.db_roles[person] = role
    rec.save_db()
    return {"ok": True, "person": person, "role": role, "people": rec.people_summary()}


@app.post("/remove-person")
async def remove_person(person: str = Form(...)):
    """Xoá 1 người khỏi DB (không cần train). Ghi lại face_db.pt."""
    person = (person or "").strip()
    ok = rec.remove_person(person)
    if ok:
        rec.save_db()
    return {"ok": ok, "person": person, "people": rec.people_summary()}


@app.get("/people/samples")
def people_samples(person: str):
    """Chẩn đoán từng mẫu của một người để tìm mẫu enroll nhầm."""
    return {"person": person, "samples": rec.sample_diagnostics((person or "").strip())}


@app.post("/remove-sample")
async def remove_sample(person: str = Form(...), index: int = Form(...)):
    """Xoá một mẫu cụ thể của 'person' theo index lấy từ /people/samples. Ghi lại face_db.pt."""
    person = (person or "").strip()
    ok = rec.remove_embedding(person, index)
    if ok:
        rec.save_db()
    return {"ok": ok, "person": person, "people": rec.people_summary()}

def _room_snapshot():
    """Snapshot trạng thái phòng theo vai trò dựa trên _in_room hiện tại."""
    ks_present = [n for n in _in_room if rec.db_roles.get(n, "kỹ sư") == "kỹ sư"]
    sv_present = [n for n in _in_room if rec.db_roles.get(n, "sinh viên") == "sinh viên"]
    total_ks   = sum(1 for n in rec.db_names if rec.db_roles.get(n, "kỹ sư") == "kỹ sư")
    total_sv   = sum(1 for n in rec.db_names if rec.db_roles.get(n, "sinh viên") == "sinh viên")
    return ks_present, sv_present, total_ks, total_sv


def _log_attendance(person: str, cam: str):
    """Check-in/out theo camera (xem CAM_ROLE), gọi từ _log_scan() khi một track được
    xác nhận OK. OK lặp lại khi di chuyển không tạo thêm sự kiện vì đã kiểm tra _in_room
    trước khi ghi."""
    role = CAM_ROLE.get(cam)
    now = datetime.now()
    if role == "vào" and person not in _in_room:
        _in_room[person] = now
        ks_p, sv_p, t_ks, t_sv = _room_snapshot()
        _attendance_log.append({
            "dt": now, "type": "vào", "person": person,
            "role": rec.db_roles.get(person, "kỹ sư"),
            "duration_min": None,
            "ks_present": list(ks_p), "sv_present": list(sv_p),
            "total_ks": t_ks, "total_sv": t_sv,
        })
    elif role == "ra" and person in _in_room:
        entry_time = _in_room.pop(person)
        dur = round((now - entry_time).total_seconds() / 60, 1)
        ks_p, sv_p, t_ks, t_sv = _room_snapshot()
        _attendance_log.append({
            "dt": now, "type": "ra", "person": person,
            "role": rec.db_roles.get(person, "kỹ sư"),
            "duration_min": dur,
            "ks_present": list(ks_p), "sv_present": list(sv_p),
            "total_ks": t_ks, "total_sv": t_sv,
        })


def _sessions_from_events(events: list[dict]) -> list[dict]:
    """Ghép cặp vào/ra thành session. Trả về list {person, role, date, entry_dt,
    duration_min, ks_present, sv_present}."""
    pending  = {}
    sessions = []
    for ev in sorted(events, key=lambda e: e["dt"]):
        if ev["type"] == "vào":
            pending[ev["person"]] = ev
        elif ev["type"] == "ra" and ev["person"] in pending:
            vao = pending.pop(ev["person"])
            sessions.append({"person": ev["person"], "role": ev.get("role", "kỹ sư"),
                              "date": vao["dt"].date(), "entry_dt": vao["dt"],
                              "duration_min": ev["duration_min"],
                              "ks_present": vao.get("ks_present", []),
                              "sv_present": vao.get("sv_present", [])})
    for person, vao in pending.items():
        sessions.append({"person": person, "role": vao.get("role", "kỹ sư"),
                         "date": vao["dt"].date(), "entry_dt": vao["dt"],
                         "duration_min": None,
                         "ks_present": vao.get("ks_present", []),
                         "sv_present": vao.get("sv_present", [])})
    return sorted(sessions, key=lambda s: s["entry_dt"])


def _make_excel(events: list[dict], scans: list[dict], people: list[dict],
                mode: str = "full") -> bytes:
    """
    Xuất 3 sheet: Ngày / Tuần / Tháng.
    - mode="full": nhật ký đầy đủ (đứng im + di chuyển + fail/lạ).
    - mode="once": chỉ nhận dạng mỗi người 1 lần (bỏ di chuyển, bỏ lặp, bỏ fail/lạ).
    """
    from openpyxl.utils import get_column_letter

    if mode == "once":
        # chỉ giữ lần OK đầu tiên của mỗi người
        seen, filtered = set(), []
        for e in sorted(scans, key=lambda x: x["dt"]):
            if e["status"] == "OK" and e["person"] not in seen:
                seen.add(e["person"]); filtered.append(e)
        scans = filtered

    now           = datetime.now()
    today         = now.date()
    week_start    = today - timedelta(days=today.weekday())
    month_start   = today.replace(day=1)
    days_in_month = calendar.monthrange(today.year, today.month)[1]

    VIET_DAYS = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]

    # Định dạng ô
    TITLE_KS   = PatternFill("solid", fgColor="1E3A5F")
    TITLE_SV   = PatternFill("solid", fgColor="4A235A")
    HDR_KS     = PatternFill("solid", fgColor="2874A6")
    HDR_SV     = PatternFill("solid", fgColor="7D3C98")
    DATE_FILL  = PatternFill("solid", fgColor="2C3E50")
    SUM_KS     = PatternFill("solid", fgColor="1A5276")
    SUM_SV     = PatternFill("solid", fgColor="6C3483")
    ALT_KS     = PatternFill("solid", fgColor="EBF5FB")
    ALT_SV     = PatternFill("solid", fgColor="F5EEF8")
    GREEN_DARK = PatternFill("solid", fgColor="1E8449")
    RED_DARK   = PatternFill("solid", fgColor="C0392B")
    GREEN_ROW  = PatternFill("solid", fgColor="D5F5E3")
    RED_ROW    = PatternFill("solid", fgColor="FADBD8")

    TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    DATE_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SUM_FONT   = Font(bold=True, color="FFFFFF", size=10)
    NORM       = Font(size=10)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")
    THIN   = Side(style="thin", color="BDBDBD")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Danh sách người theo vai trò
    ks_names  = sorted([p["name"] for p in people if p.get("role") == "kỹ sư"])
    sv_names  = sorted([p["name"] for p in people if p.get("role") == "sinh viên"])
    total_ks  = len(ks_names)
    total_sv  = len(sv_names)

    # Ghép session và tổng hợp
    all_sessions = _sessions_from_events(events)

    def _agg_daily(sessions):
        """date -> person -> {role, first_entry, last_exit, total_min, count}"""
        agg = defaultdict(dict)
        for s in sessions:
            d, p = s["date"], s["person"]
            if p not in agg[d]:
                agg[d][p] = {"role": s["role"], "first_entry": s["entry_dt"],
                              "last_exit": None, "total_min": 0.0, "count": 0}
            r = agg[d][p]
            if s["entry_dt"] < r["first_entry"]:
                r["first_entry"] = s["entry_dt"]
            if s["duration_min"] is not None:
                exit_dt = s["entry_dt"] + timedelta(minutes=s["duration_min"])
                if r["last_exit"] is None or exit_dt > r["last_exit"]:
                    r["last_exit"] = exit_dt
                r["total_min"] += s["duration_min"]
            r["count"] += 1
        return agg

    daily_agg = _agg_daily(all_sessions)
    all_dates = sorted(daily_agg.keys())

    def _fmt_dur(minutes):
        if not minutes:
            return "—"
        h, m = int(minutes // 60), int(minutes % 60)
        return f"{h}h{m:02d}m" if h else f"{m}m"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Ngày")
    ORANGE_ROW = PatternFill("solid", fgColor="FDEBD0")

    OK_COLS   = ["STT", "Ngày", "Giờ", "Họ và Tên", "Camera", "Trạng thái"]
    OK_WIDTHS = [5, 10, 9, 18, 14, 11]
    NC  = len(OK_COLS)
    KS1 = 1
    GAP = KS1 + NC            # cột giữa 2 bảng (cũng là cột "Kết quả" của bảng fail)
    SV1 = GAP + 1

    # tiêu đề 2 bảng
    ws1.merge_cells(start_row=1, start_column=KS1, end_row=1, end_column=KS1+NC-1)
    c = ws1.cell(row=1, column=KS1, value="BẢNG NHÂN VIÊN (nhận diện OK)")
    c.font = TITLE_FONT; c.fill = TITLE_KS; c.alignment = CENTER
    ws1.merge_cells(start_row=1, start_column=SV1, end_row=1, end_column=SV1+NC-1)
    c = ws1.cell(row=1, column=SV1, value="BẢNG SINH VIÊN  (nhận diện OK)")
    c.font = TITLE_FONT; c.fill = TITLE_SV; c.alignment = CENTER
    ws1.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(OK_COLS, OK_WIDTHS)):
        for start, fill in [(KS1, HDR_KS), (SV1, HDR_SV)]:
            c = ws1.cell(row=2, column=start+ci, value=h)
            c.font = HDR_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER
        ws1.column_dimensions[get_column_letter(KS1+ci)].width = w
        ws1.column_dimensions[get_column_letter(SV1+ci)].width = w
    ws1.column_dimensions[get_column_letter(GAP)].width = 11   # gap + cột Kết quả bảng fail
    ws1.row_dimensions[2].height = 22
    ws1.freeze_panes = "A3"

    # tách sự kiện OK theo vai trò
    ok_ks = sorted([e for e in scans if e["status"] == "OK" and e["role"] == "kỹ sư"],
                   key=lambda e: e["dt"])
    ok_sv = sorted([e for e in scans if e["status"] == "OK" and e["role"] == "sinh viên"],
                   key=lambda e: e["dt"])

    def _fill_ok(events, base, alt_fill):
        for i, e in enumerate(events):
            ri = 3 + i
            vals = [i+1, e["dt"].strftime("%d/%m"), e["dt"].strftime("%H:%M:%S"),
                    e["person"], e.get("cam", ""), e.get("motion", "Đứng im")]
            for ci, val in enumerate(vals):
                c = ws1.cell(row=ri, column=base+ci, value=val)
                c.border = BORDER; c.font = NORM
                c.alignment = LEFT if ci == 3 else CENTER
                if i % 2 == 1: c.fill = alt_fill
            ws1.row_dimensions[ri].height = 20

    _fill_ok(ok_ks, KS1, ALT_KS)
    _fill_ok(ok_sv, SV1, ALT_SV)

    n_rows  = max(len(ok_ks), len(ok_sv), 1)
    sum_row = 3 + n_rows
    ws1.merge_cells(start_row=sum_row, start_column=KS1, end_row=sum_row, end_column=KS1+NC-1)
    c = ws1.cell(row=sum_row, column=KS1,
                 value=f"Nhân viên: {len(ok_ks)} lần OK  ({len({e['person'] for e in ok_ks})} người)")
    c.font = SUM_FONT; c.fill = SUM_KS; c.alignment = CENTER
    ws1.merge_cells(start_row=sum_row, start_column=SV1, end_row=sum_row, end_column=SV1+NC-1)
    c = ws1.cell(row=sum_row, column=SV1,
                 value=f"Sinh viên: {len(ok_sv)} lần OK  ({len({e['person'] for e in ok_sv})} người)")
    c.font = SUM_FONT; c.fill = SUM_SV; c.alignment = CENTER
    ws1.row_dimensions[sum_row].height = 22

    # Bảng người lạ / fail (bên dưới, trải hết chiều ngang)
    ftitle = sum_row + 2
    ws1.merge_cells(start_row=ftitle, start_column=1, end_row=ftitle, end_column=SV1+NC-1)
    c = ws1.cell(row=ftitle, column=1, value="BẢNG NGƯỜI LẠ / NHẬN DIỆN FAIL")
    c.font = TITLE_FONT; c.fill = RED_DARK; c.alignment = CENTER
    ws1.row_dimensions[ftitle].height = 26

    fhdr = ftitle + 1
    for ci, h in enumerate(["STT", "Ngày", "Giờ", "Họ và Tên", "Camera", "Trạng thái", "Kết quả"]):
        c = ws1.cell(row=fhdr, column=1+ci, value=h)
        c.font = HDR_FONT; c.fill = RED_DARK; c.alignment = CENTER; c.border = BORDER
    ws1.row_dimensions[fhdr].height = 22

    fails = sorted([e for e in scans if e["status"] in ("FAIL", "LẠ")], key=lambda e: e["dt"])
    for i, e in enumerate(fails):
        ri = fhdr + 1 + i
        fill = RED_ROW if e["status"] == "FAIL" else ORANGE_ROW
        fcol = "C0392B" if e["status"] == "FAIL" else "CA6F1E"
        vals = [i+1, e["dt"].strftime("%d/%m"), e["dt"].strftime("%H:%M:%S"),
                e["person"], e.get("cam", ""), e.get("motion", "Đứng im"), e["status"]]
        for ci, val in enumerate(vals):
            c = ws1.cell(row=ri, column=1+ci, value=val)
            c.border = BORDER; c.fill = fill
            c.alignment = LEFT if ci == 3 else CENTER
            c.font = Font(bold=True, size=10, color=fcol) if ci == 6 else NORM
        ws1.row_dimensions[ri].height = 20
    if not fails:
        c = ws1.cell(row=fhdr+1, column=1, value="Không có fail / người lạ.")
        c.font = NORM; c.alignment = LEFT

    W_COLS   = ["STT", "Họ và Tên", "Ngày đi làm"]
    W_WIDTHS = [5, 22, 15]
    W_NCOLS  = len(W_COLS)
    KS2 = 1
    SV2 = KS2 + W_NCOLS + 1

    ws2 = wb.create_sheet("Tuần")

    ws2.merge_cells(start_row=1, start_column=KS2, end_row=1, end_column=KS2+W_NCOLS-1)
    c = ws2.cell(row=1, column=KS2, value="BẢNG NHÂN VIÊN")
    c.font = TITLE_FONT; c.fill = TITLE_KS; c.alignment = CENTER

    ws2.merge_cells(start_row=1, start_column=SV2, end_row=1, end_column=SV2+W_NCOLS-1)
    c = ws2.cell(row=1, column=SV2, value="BẢNG SINH VIÊN")
    c.font = TITLE_FONT; c.fill = TITLE_SV; c.alignment = CENTER
    ws2.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(W_COLS, W_WIDTHS)):
        for start, fill in [(KS2, HDR_KS), (SV2, HDR_SV)]:
            c = ws2.cell(row=2, column=start+ci, value=h)
            c.font = HDR_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER
        ws2.column_dimensions[get_column_letter(KS2+ci)].width = w
        ws2.column_dimensions[get_column_letter(SV2+ci)].width = w
    ws2.column_dimensions[get_column_letter(KS2+W_NCOLS)].width = 2
    ws2.row_dimensions[2].height = 22
    ws2.freeze_panes = "B3"

    # Gom số liệu theo tuần
    week_to_dates_ks = defaultdict(set)
    week_to_dates_sv = defaultdict(set)
    for s in all_sessions:
        ws = s["date"] - timedelta(days=s["date"].weekday())
        if s["role"] == "kỹ sư":
            week_to_dates_ks[(ws, s["person"])].add(s["date"])
        else:
            week_to_dates_sv[(ws, s["person"])].add(s["date"])

    all_weeks = sorted(set([s["date"] - timedelta(days=s["date"].weekday()) for s in all_sessions]))

    cur_row = 3
    for ws in all_weeks:
        we = ws + timedelta(days=6)
        days_in_week = (min(we, today) - ws).days + 1 if ws <= today else 7
        week_label = f"  Tuần {ws.strftime('%d/%m')} – {we.strftime('%d/%m/%Y')}"

        ws2.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=SV2+W_NCOLS-1)
        c = ws2.cell(row=cur_row, column=1, value=week_label)
        c.font = DATE_FONT; c.fill = DATE_FILL; c.alignment = LEFT
        ws2.row_dimensions[cur_row].height = 24
        cur_row += 1

        max_people = max(len(ks_names), len(sv_names), 1)
        for i in range(max_people):
            ri = cur_row + i
            alt = i % 2 == 1
            ws2.row_dimensions[ri].height = 20

            if i < len(ks_names):
                name = ks_names[i]
                cnt = len(week_to_dates_ks[(ws, name)])
                for ci, val in enumerate([i+1, name, f"{cnt}/{days_in_week} ngày"]):
                    c = ws2.cell(row=ri, column=KS2+ci, value=val)
                    c.border = BORDER; c.font = NORM
                    c.alignment = LEFT if ci == 1 else CENTER
                    if alt: c.fill = ALT_KS

            if i < len(sv_names):
                name = sv_names[i]
                cnt = len(week_to_dates_sv[(ws, name)])
                for ci, val in enumerate([i+1, name, f"{cnt}/{days_in_week} ngày"]):
                    c = ws2.cell(row=ri, column=SV2+ci, value=val)
                    c.border = BORDER; c.font = NORM
                    c.alignment = LEFT if ci == 1 else CENTER
                    if alt: c.fill = ALT_SV

        cur_row += max_people

        ks_present_w = len([n for n in ks_names if week_to_dates_ks[(ws, n)]])
        sv_present_w = len([n for n in sv_names if week_to_dates_sv[(ws, n)]])

        ws2.merge_cells(start_row=cur_row, start_column=KS2, end_row=cur_row, end_column=KS2+W_NCOLS-1)
        c = ws2.cell(row=cur_row, column=KS2, value=f"Nhân viên đi làm: {ks_present_w}/{total_ks} người")
        c.font = SUM_FONT; c.fill = SUM_KS; c.alignment = CENTER

        ws2.merge_cells(start_row=cur_row, start_column=SV2, end_row=cur_row, end_column=SV2+W_NCOLS-1)
        c = ws2.cell(row=cur_row, column=SV2, value=f"Sinh viên đi làm: {sv_present_w}/{total_sv} người")
        c.font = SUM_FONT; c.fill = SUM_SV; c.alignment = CENTER
        ws2.row_dimensions[cur_row].height = 22
        cur_row += 2

    M_COLS   = ["STT", "Họ và Tên", "Ngày đi làm"]
    M_WIDTHS = [5, 22, 18]
    M_NCOLS  = len(M_COLS)
    KS3 = 1
    SV3 = KS3 + M_NCOLS + 1

    ws3 = wb.create_sheet("Tháng")

    ws3.merge_cells(start_row=1, start_column=KS3, end_row=1, end_column=KS3+M_NCOLS-1)
    c = ws3.cell(row=1, column=KS3, value="BẢNG NHÂN VIÊN")
    c.font = TITLE_FONT; c.fill = TITLE_KS; c.alignment = CENTER

    ws3.merge_cells(start_row=1, start_column=SV3, end_row=1, end_column=SV3+M_NCOLS-1)
    c = ws3.cell(row=1, column=SV3, value="BẢNG SINH VIÊN")
    c.font = TITLE_FONT; c.fill = TITLE_SV; c.alignment = CENTER
    ws3.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(M_COLS, M_WIDTHS)):
        for start, fill in [(KS3, HDR_KS), (SV3, HDR_SV)]:
            c = ws3.cell(row=2, column=start+ci, value=h)
            c.font = HDR_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER
        ws3.column_dimensions[get_column_letter(KS3+ci)].width = w
        ws3.column_dimensions[get_column_letter(SV3+ci)].width = w
    ws3.column_dimensions[get_column_letter(KS3+M_NCOLS)].width = 2
    ws3.row_dimensions[2].height = 22
    ws3.freeze_panes = "B3"

    # Gom số liệu theo tháng
    month_to_dates_ks = defaultdict(set)
    month_to_dates_sv = defaultdict(set)
    for s in all_sessions:
        ms = (s["date"].year, s["date"].month)
        if s["role"] == "kỹ sư":
            month_to_dates_ks[(ms, s["person"])].add(s["date"])
        else:
            month_to_dates_sv[(ms, s["person"])].add(s["date"])

    all_months = sorted(set([(s["date"].year, s["date"].month) for s in all_sessions]))

    cur_row = 3
    for ym in all_months:
        y, m = ym
        month_start_dt = datetime(y, m, 1).date()
        month_end_dt = datetime(y, m, calendar.monthrange(y, m)[1]).date()
        days_in_m = calendar.monthrange(y, m)[1]
        month_label = f"  Tháng {m}/{y} ({days_in_m} ngày)"

        ws3.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=SV3+M_NCOLS-1)
        c = ws3.cell(row=cur_row, column=1, value=month_label)
        c.font = DATE_FONT; c.fill = DATE_FILL; c.alignment = LEFT
        ws3.row_dimensions[cur_row].height = 24
        cur_row += 1

        max_people = max(len(ks_names), len(sv_names), 1)
        for i in range(max_people):
            ri = cur_row + i
            alt = i % 2 == 1
            ws3.row_dimensions[ri].height = 20

            if i < len(ks_names):
                name = ks_names[i]
                cnt = len(month_to_dates_ks[(ym, name)])
                for ci, val in enumerate([i+1, name, f"{cnt}/{days_in_m} ngày"]):
                    c = ws3.cell(row=ri, column=KS3+ci, value=val)
                    c.border = BORDER; c.font = NORM
                    c.alignment = LEFT if ci == 1 else CENTER
                    if alt: c.fill = ALT_KS

            if i < len(sv_names):
                name = sv_names[i]
                cnt = len(month_to_dates_sv[(ym, name)])
                for ci, val in enumerate([i+1, name, f"{cnt}/{days_in_m} ngày"]):
                    c = ws3.cell(row=ri, column=SV3+ci, value=val)
                    c.border = BORDER; c.font = NORM
                    c.alignment = LEFT if ci == 1 else CENTER
                    if alt: c.fill = ALT_SV

        cur_row += max_people

        ks_present_m = len([n for n in ks_names if month_to_dates_ks[(ym, n)]])
        sv_present_m = len([n for n in sv_names if month_to_dates_sv[(ym, n)]])

        ws3.merge_cells(start_row=cur_row, start_column=KS3, end_row=cur_row, end_column=KS3+M_NCOLS-1)
        c = ws3.cell(row=cur_row, column=KS3, value=f"Nhân viên đi làm: {ks_present_m}/{total_ks} người")
        c.font = SUM_FONT; c.fill = SUM_KS; c.alignment = CENTER

        ws3.merge_cells(start_row=cur_row, start_column=SV3, end_row=cur_row, end_column=SV3+M_NCOLS-1)
        c = ws3.cell(row=cur_row, column=SV3, value=f"Sinh viên đi làm: {sv_present_m}/{total_sv} người")
        c.font = SUM_FONT; c.fill = SUM_SV; c.alignment = CENTER
        ws3.row_dimensions[cur_row].height = 22
        cur_row += 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@app.get("/export")
def export_attendance(mode: str = "full"):
    """Xuất nhật ký ra Excel. mode: 'full' (đầy đủ có di chuyển) | 'once' (mỗi người 1 lần)."""
    mode = "once" if mode == "once" else "full"
    data = _make_excel(_attendance_log, _scan_log, rec.people_summary(), mode=mode)
    tag = "1lan" if mode == "once" else "daydu"
    filename = f"nhat_ky_{tag}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# Các ảnh biểu đồ mà YOLO xuất ra sau khi train.
PLOT_FILES = [
    "confusion_matrix_normalized.png", "confusion_matrix.png",
    "results.png", "BoxPR_curve.png", "BoxF1_curve.png",
    "val_batch0_pred.jpg", "val_batch0_labels.jpg",
]


def find_runs():
    """Tìm mọi thư mục run (chứa results.png) dưới BASE_DIR/runs. Trả {tên: đường_dẫn}."""
    runs = {}
    for p in glob.glob(os.path.join(BASE_DIR, "runs", "**", "results.png"), recursive=True):
        d = os.path.dirname(p)
        runs[os.path.basename(d)] = d
    # mới nhất trước
    return dict(sorted(runs.items(), key=lambda kv: os.path.getmtime(kv[1]), reverse=True))


@app.get("/results/file")
def results_file(run: str, f: str):
    """1 ảnh biểu đồ của 1 run."""
    runs = find_runs()
    d = runs.get(run)
    if not d or f not in PLOT_FILES:
        return JSONResponse({"error": "not found"}, status_code=404)
    fp = os.path.join(d, f)
    if not os.path.isfile(fp):
        return JSONResponse({"error": "no file"}, status_code=404)
    return FileResponse(fp)


@app.get("/results", response_class=HTMLResponse)
def results_page(run: str = ""):
    """Trang hiển thị confusion matrix, mAP curve, ảnh predict của run đã train."""
    runs = find_runs()
    if not runs:
        return "<body style='background:#0f172a;color:#e2e8f0;font-family:Arial;padding:40px'>" \
               "<h2>Chưa có run nào</h2>" \
               "<a href='/' style='color:#38bdf8'>← Về trang nhận diện</a></body>"
    if run not in runs:
        run = next(iter(runs))  # mặc định: run mới nhất

    options = "".join(
        f"<option value='{n}'{' selected' if n == run else ''}>{n}</option>" for n in runs
    )
    titles = {
        "confusion_matrix_normalized.png": "Confusion matrix (chuẩn hoá)",
        "results.png": "Đường cong loss và mAP theo epoch",
        "BoxPR_curve.png": "Precision - Recall",
        "BoxF1_curve.png": "F1 theo confidence",
        "val_batch0_pred.jpg": "Ảnh dự đoán (val)",
        "val_batch0_labels.jpg": "Ảnh nhãn gốc (val)",
    }
    cards = ""
    for f, title in titles.items():
        if os.path.isfile(os.path.join(runs[run], f)):
            cards += (
                f"<div class='card'><h3>{title}</h3>"
                f"<a href='/results/file?run={run}&f={f}' target='_blank'>"
                f"<img src='/results/file?run={run}&f={f}'></a></div>"
            )
    return f"""<!DOCTYPE html><html lang='vi'><head><meta charset='UTF-8'>
<meta name='viewport' content='width=device-width, initial-scale=1.0'>
<title>Kết quả Train</title><style>
*{{box-sizing:border-box}} body{{background:#0f172a;color:#e2e8f0;font-family:Arial;margin:0;padding:20px}}
h1{{color:#38bdf8;text-align:center;margin:0 0 4px}} .sub{{text-align:center;color:#94a3b8;margin-bottom:16px}}
.top{{max-width:1100px;margin:0 auto 18px;display:flex;gap:12px;align-items:center;justify-content:center;flex-wrap:wrap}}
select{{padding:9px;background:#334155;color:#fff;border:1px solid #475569;border-radius:6px;font-size:14px}}
a.back{{color:#38bdf8;text-decoration:none}}
.grid{{max-width:1100px;margin:0 auto;display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:16px}}
.card{{background:#1e293b;border:1px solid #334155;border-radius:12px;padding:14px}}
.card h3{{margin:0 0 10px;font-size:14px;color:#cbd5e1}}
.card img{{width:100%;border-radius:8px;background:#fff}}
</style></head><body>
<h1>Kết quả Training</h1><div class='sub'>Run: <b>{run}</b></div>
<div class='top'>
  <form method='get' action='/results'>
    <label style='color:#94a3b8;font-size:14px'>Chọn run: </label>
    <select name='run' onchange='this.form.submit()'>{options}</select>
  </form>
  <a class='back' href='/'>← Về trang nhận diện</a>
</div>
<div class='grid'>{cards or "<p>Run này chưa có ảnh biểu đồ.</p>"}</div>
</body></html>"""


